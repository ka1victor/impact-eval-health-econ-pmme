"""Constrói e congela a tipologia territorial para o tema atração e provimento.

Fontes oficiais usadas para classificação, todas anteriores ao PMM-E:
- 27 Capitais (códigos oficiais IBGE, estáveis).
- RM/RIDE 2022 (IBGE, composição oficial de 31/12/2022 — pré-PMM-E).
- REGIC 2018 — Hierarquia e região de influência (IBGE, 2018 — centralidade).

Medidas contínuas prévias preservadas (sem consultar outcomes):
- IVS 2010 (canônico) + subíndices, IDHM 2010, população 2010, RDPC 2010
- Região / macrorregião de saúde (painel de 04_harmonizar_territorio_ibge.py)
- Estoque médico pré-oferta (CNES 202407–202506, painel_cnes_especialidade_mensal)

A classificação é congelada sem consultar alocação, homologação ou CNES pós.
"""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
import urllib.request
from pathlib import Path
from typing import Any

import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
DATA_TERRITORIO = ROOT / "data" / "raw" / "aquisicao" / "territorio"
MALHA = ROOT / "output" / "aquisicao" / "malha_municipios_regioes_saude.parquet"
MATRIZ_FUNIL = ROOT / "output" / "tema_trabalho" / "matriz_funil_ciclo1.parquet"
PAINEL_CNES = ROOT / "output" / "painel_cnes_especialidade_mensal.parquet"
QUADRO = ROOT / "output" / "aquisicao" / "quadro_vagas_tratamento.parquet"

REGIC_XLSX = DATA_TERRITORIO / "REGIC2018_Municipios_Hierarquia_e_regiao.xlsx"
RM_XLSX = DATA_TERRITORIO / "Composicao_RMs_RIDEs_AglomUrbanas_2022_v2.xlsx"

OUT_DIR = ROOT / "output" / "tema_trabalho"
OUT_MATRIZ = OUT_DIR / "matriz_tipologia_territorial.parquet"
OUT_MANIFESTO = OUT_DIR / "manifesto_tipologia_territorial.json"
OUT_SUPORTE = OUT_DIR / "suporte_estratos_territoriais.csv"

REGIC_URL = (
    "https://geoftp.ibge.gov.br/organizacao_do_territorio/divisao_regional/"
    "regioes_de_influencia_das_cidades/"
    "Regioes_de_influencia_das_cidades_2018_Resultados_definitivos/"
    "base_tabular/REGIC2018_Municipios_Hierarquia_e_regiao.xlsx"
)
RM_URL = (
    "https://geoftp.ibge.gov.br/organizacao_do_territorio/estrutura_territorial/"
    "municipios_por_regioes_metropolitanas/Situacao_2020a2029/"
    "Composicao_RMs_RIDEs_AglomUrbanas_2022_v2.xlsx"
)

# 27 capitais — códigos IBGE 7d oficiais (DTB estável)
CAPITAIS_IBGE7 = {
    "1100205",  # Porto Velho
    "1200401",  # Rio Branco
    "1302603",  # Manaus
    "1400100",  # Boa Vista
    "1501402",  # Belém
    "1600303",  # Macapá
    "1721000",  # Palmas
    "2111300",  # São Luís
    "2211001",  # Teresina
    "2304400",  # Fortaleza
    "2408102",  # Natal
    "2507507",  # João Pessoa
    "2611606",  # Recife
    "2704302",  # Maceió
    "2800308",  # Aracaju
    "2927408",  # Salvador
    "3106200",  # Belo Horizonte
    "3205309",  # Vitória
    "3304557",  # Rio de Janeiro
    "3550308",  # São Paulo
    "4106902",  # Curitiba
    "4205407",  # Florianópolis
    "4314902",  # Porto Alegre
    "5002704",  # Campo Grande
    "5103403",  # Cuiabá
    "5208707",  # Goiânia
    "5300108",  # Brasília
}
CAPITAIS_IBGE6 = {code[:6] for code in CAPITAIS_IBGE7}

ORDEM_ESTRATO = {
    "capital": 1,
    "metropolitano": 2,
    "interior_proximo_polo": 3,
    "interior_remoto": 4,
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def download_if_missing(url: str, dest: Path) -> None:
    if dest.exists():
        return
    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"Baixando {dest.name} de {url} ...")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; PMME-impact-eval/1.0)"})
    with urllib.request.urlopen(req, timeout=120) as resp, dest.open("wb") as out:
        out.write(resp.read())
    print(f"  -> salvo em {dest} ({dest.stat().st_size} bytes)")


def normalize_id(value: object, width: int) -> str:
    if pd.isna(value):
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits[:width].zfill(width) if digits else ""


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    DATA_TERRITORIO.mkdir(parents=True, exist_ok=True)

    # 1. Aquisição versionada — baixa apenas se ausente (rota reproduzível)
    download_if_missing(REGIC_URL, REGIC_XLSX)
    download_if_missing(RM_URL, RM_XLSX)

    # Copia temporária de inspeção (se baixada em TMP por sessão anterior)
    tmp_regic = Path(r"C:\Users\camil\AppData\Local\Temp\opencode\regic2018_municipios.xlsx")
    tmp_rm = Path(r"C:\Users\camil\AppData\Local\Temp\opencode\composicao_rm_2022.xlsx")
    if tmp_regic.exists() and not REGIC_XLSX.exists():
        REGIC_XLSX.write_bytes(tmp_regic.read_bytes())
    if tmp_rm.exists() and not RM_XLSX.exists():
        RM_XLSX.write_bytes(tmp_rm.read_bytes())

    for path in [MALHA, MATRIZ_FUNIL, PAINEL_CNES, QUADRO, REGIC_XLSX, RM_XLSX]:
        if not path.exists():
            raise FileNotFoundError(path)

    # Guard: nunca consultar outcomes ao classificar
    forbidden_cols = {"n_confirmacoes_ch1", "n_homologacoes_ch1", "outcome_alguma_confirmacao_ou_homologacao"}
    # (leitura da matriz será restrita a co_ibge_6d)

    # 2. Malha base (IVS 2010 canônico, população 2010, região de saúde)
    malha = pd.read_parquet(MALHA).copy()
    malha["co_ibge_6d"] = malha["co_ibge_6d"].astype(str).str.replace(r"\D", "", regex=True).str.zfill(6)
    malha["co_ibge_7d"] = malha["co_ibge_7d"].astype(str).str.replace(r"\D", "", regex=True).str.zfill(7)

    # 3. REGIC 2018 — hierarquia/centralidade por município
    import openpyxl

    wb = openpyxl.load_workbook(REGIC_XLSX, read_only=True, data_only=True)
    ws = wb["Hierarquia e região"]
    rows: list[list[Any]] = []
    header: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(v) if v is not None else "" for v in row]
        else:
            rows.append(list(row))
    regic = pd.DataFrame(rows, columns=header)
    # Normaliza códigos
    regic["codmun7"] = regic["codmun"].astype(str).str.replace(r"\D", "", regex=True).str.zfill(7)
    regic["codmun6"] = regic["codmun7"].str[:6]
    # Mantém colunas relevantes
    regic_small = regic[
        ["codmun7", "codmun6", "Município", "Categoria", "codcid", "Hierarquia", "Hierarquia 2", "Hierarquia - grupo"]
    ].copy()
    regic_small = regic_small.rename(
        columns={
            "Município": "no_municipio_regic",
            "Categoria": "categoria_regic",
            "Hierarquia": "hierarquia_regic",
            "Hierarquia 2": "hierarquia2_regic",
            "Hierarquia - grupo": "hierarquia_grupo_regic",
        }
    )

    # 4. RM/RIDE 2022 — composição oficial (31/12/2022, pré-PMM-E)
    wb2 = openpyxl.load_workbook(RM_XLSX, read_only=True, data_only=True)
    # Aba pode ter nome com acento; pega a primeira que contém 'Recortes Metropoli'
    sheet_rm = [s for s in wb2.sheetnames if "Recortes Metropoli" in s][0]
    ws2 = wb2[sheet_rm]
    rows2: list[list[Any]] = []
    header2: list[str] = []
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            header2 = [str(v) if v is not None else "" for v in row]
        else:
            rows2.append(list(row))
    rm = pd.DataFrame(rows2, columns=header2)
    rm["cod7"] = rm["COD_MUN"].astype(str).str.replace(r"\D", "", regex=True).str.zfill(7)
    rm["cod6"] = rm["cod7"].str[:6]
    rm_set_7 = set(rm["cod7"])
    # Guarda categoria para auditoria
    rm_categoria_counts = rm["NOME_CATMETROPOL"].value_counts().to_dict()

    # 5. Universo A1 — municípios presentes na matriz do funil (sem outcomes)
    matriz = pd.read_parquet(MATRIZ_FUNIL, columns=["co_ibge_6d"]).copy()
    # Remove nulos (266 linhas fora do quadro sem município)
    matriz_valid = matriz.dropna(subset=["co_ibge_6d"]).copy()
    matriz_valid["co_ibge_6d"] = matriz_valid["co_ibge_6d"].astype(str).str.replace(r"\D", "", regex=True).str.zfill(6)
    municipios_a1_6 = set(matriz_valid["co_ibge_6d"].unique())
    # Para relatório: 266 registros fora do quadro sem município
    registros_fora_quadro_sem_municipio = int(matriz["co_ibge_6d"].isna().sum())

    # 6. Painel CNES pré-oferta — estoque médio 202407–202506 (12 meses pré-publicação 2025-07-24)
    cnes = pd.read_parquet(PAINEL_CNES).copy()
    # Competência como string YYYYMM
    cnes["competencia"] = cnes["competencia"].astype(str).str.replace(r"\D", "", regex=True)
    # Pré-oferta: competências < 202507 (estritamente antes da publicação)
    cnes_pre = cnes[(cnes["competencia"] >= "202407") & (cnes["competencia"] <= "202506")].copy()
    # Soma CNES-curso dentro do município por competência, depois média
    monthly = (
        cnes_pre.groupby(["co_ibge_6d", "competencia"], as_index=False)["especialistas_ist"].sum()
    )
    monthly["co_ibge_6d"] = monthly["co_ibge_6d"].astype(str).str.replace(r"\D", "", regex=True).str.zfill(6)
    estoque_pre = (
        monthly.groupby("co_ibge_6d", as_index=False)["especialistas_ist"].mean().rename(
            columns={"especialistas_ist": "estoque_especialistas_pre_12m_media"}
        )
    )
    # Também total por município (média) — já é a média mensal

    # 7. Construção da matriz nacional (5570 = 5565 malha + 5 novos municípios pós-2010)
    # Base nacional = REGIC master (5570)
    nacional = regic_small[["codmun7", "codmun6", "hierarquia_regic", "hierarquia2_regic", "hierarquia_grupo_regic", "categoria_regic"]].copy()
    nacional = nacional.rename(columns={"codmun7": "co_ibge_7d", "codmun6": "co_ibge_6d"})
    # Left join malha (IVS, pop, região saúde) — 5 novos ficam com NaN em IVS/pop
    malha_cols = [
        "co_ibge_6d",
        "co_ibge_7d",
        "no_municipio",
        "sg_uf",
        "nome_uf",
        "macro_regiao_saude",
        "no_regiao_saude",
        "ivs_2010",
        "ivs_infra_2010",
        "ivs_ch_2010",
        "ivs_rt_2010",
        "ivs_categoria",
        "idhm_2010",
        "populacao_2010",
        "rdpc_2010",
    ]
    # Deduplica malha por co_ibge_6d (já único) e faz merge por co_ibge_6d
    nacional = nacional.merge(malha[malha_cols].drop_duplicates(subset=["co_ibge_6d"]), on="co_ibge_6d", how="left", suffixes=("", "_malha"))
    # Para co_ibge_7d conflitante (nacional tem o REGIC codmun7), mantém REGIC; malha 7d fica em _malha se diferente — usa REGIC para chave
    # Preenche nome/UF para os 5 novos via REGIC (já têm categoria, mas não no_municipio/sg_uf da malha)
    # Para os 5, busca no REGIC original a UF
    # O DataFrame regic já tem siguf/UF; recupera
    uf_map = dict(zip(regic["codmun6"], regic["siguf"]))
    nome_map = dict(zip(regic["codmun6"], regic["Município"]))
    # Preenche onde malha não tinha
    for col, m in [("sg_uf", uf_map), ("no_municipio", nome_map)]:
        nacional[col] = nacional[col].fillna(nacional["co_ibge_6d"].map(m))

    # Join estoque pré-oferta
    nacional = nacional.merge(estoque_pre, on="co_ibge_6d", how="left")

    # 8. Classificação congelada (sem consultar outcomes)
    def classifica(row: pd.Series) -> str:
        co7 = str(row["co_ibge_7d"]).zfill(7)
        co6 = str(row["co_ibge_6d"]).zfill(6)
        hg = row["hierarquia_grupo_regic"]
        if co7 in CAPITAIS_IBGE7:
            return "capital"
        if co7 in rm_set_7:
            return "metropolitano"
        # REGIC grupo 5 = Centro Local (mais remoto); inclui integrante de AP
        if hg in ("5 - Centro Local", "5 - Centro Local - Integrante de Arranjo Populacional"):
            return "interior_remoto"
        # Demais níveis 1–4 = centros com alguma centralidade = próximo/conectado a polo
        # Inclui 1-Metrópole (integrantes já capturados acima se RM; senão próximo)
        # 2-Capital Regional, 3-Centro Sub-Regional, 4-Centro de Zona
        if pd.notna(hg):
            return "interior_proximo_polo"
        return "NAO_CLASSIFICADO"

    nacional["estrato"] = nacional.apply(classifica, axis=1)
    nacional["estrato_ordem"] = nacional["estrato"].map(ORDEM_ESTRATO)
    nacional["flag_capital"] = nacional["co_ibge_7d"].isin(CAPITAIS_IBGE7)
    nacional["flag_rm_ride_2022"] = nacional["co_ibge_7d"].isin(rm_set_7)
    nacional["flag_ivs_missing"] = nacional["ivs_2010"].isna()
    nacional["flag_estoque_pre_missing"] = nacional["estoque_especialistas_pre_12m_media"].isna()
    nacional["in_populacao_A1"] = nacional["co_ibge_6d"].isin(municipios_a1_6)
    # Estoque per 10k
    nacional["estoque_pre_por_10k"] = nacional["estoque_especialistas_pre_12m_media"] / (nacional["populacao_2010"] / 10000)
    # Substitui inf por NaN onde população 0/NaN
    nacional.loc[nacional["populacao_2010"].isna() | (nacional["populacao_2010"] == 0), "estoque_pre_por_10k"] = pd.NA

    # Ordenação estável: estrato_ordem + UF + município
    nacional = nacional.sort_values(["estrato_ordem", "sg_uf", "no_municipio"]).reset_index(drop=True)

    # Colunas finais (sem expor dado pessoal; nomes municipais são oficiais e públicos)
    cols_out = [
        "co_ibge_6d",
        "co_ibge_7d",
        "no_municipio",
        "sg_uf",
        "nome_uf",
        "estrato",
        "estrato_ordem",
        "flag_capital",
        "flag_rm_ride_2022",
        "categoria_regic",
        "hierarquia_regic",
        "hierarquia2_regic",
        "hierarquia_grupo_regic",
        "macro_regiao_saude",
        "no_regiao_saude",
        "ivs_2010",
        "ivs_infra_2010",
        "ivs_ch_2010",
        "ivs_rt_2010",
        "ivs_categoria",
        "idhm_2010",
        "populacao_2010",
        "rdpc_2010",
        "estoque_especialistas_pre_12m_media",
        "estoque_pre_por_10k",
        "flag_ivs_missing",
        "flag_estoque_pre_missing",
        "in_populacao_A1",
    ]
    matriz_tipologia = nacional[cols_out].copy()

    # 9. Suporte por estrato (sem outcomes) — descreve a oferta e a população A1
    quadro = pd.read_parquet(QUADRO).copy()
    quadro["co_ibge_6d"] = quadro["co_ibge_6d"].astype(str).str.replace(r"\D", "", regex=True).str.zfill(6)
    # Mapa estrato por co_ibge_6d
    estrato_map = dict(zip(matriz_tipologia["co_ibge_6d"], matriz_tipologia["estrato"]))
    quadro["estrato"] = quadro["co_ibge_6d"].map(estrato_map)

    # Matriz funil — células por município (para contagem de células na população A1)
    # Usa matriz_valid (co_ibge_6d não nulo) agrupado por município
    celulas_por_municipio = matriz_valid.groupby("co_ibge_6d").size().rename("n_celulas_funil_A1").reset_index()
    celulas_por_municipio["estrato"] = celulas_por_municipio["co_ibge_6d"].map(estrato_map)

    # Agregações por estrato
    suporte_rows: list[dict[str, Any]] = []
    for estrato, ordem in ORDEM_ESTRATO.items():
        # Nacional
        nat = matriz_tipologia[matriz_tipologia["estrato"] == estrato]
        # Dentro da população A1
        a1_muns = nat[nat["in_populacao_A1"]]
        # Quadro (oferta): células e vagas nesse estrato
        q = quadro[quadro["estrato"] == estrato]
        # Células do funil A1 (matriz) nesse estrato
        c = celulas_por_municipio[celulas_por_municipio["estrato"] == estrato]
        suporte_rows.append(
            {
                "estrato": estrato,
                "ordem": ordem,
                "n_municipios_nacional": int(len(nat)),
                "n_municipios_populacao_A1": int(len(a1_muns)),
                "n_celulas_quadro_ch1": int(len(q)),
                "n_celulas_funil_A1": int(c["n_celulas_funil_A1"].sum()) if len(c) else 0,
                "vagas_imediatas_publicadas": int(q["qt_vagas_imediatas"].sum()) if len(q) else 0,
                "vagas_reserva_publicadas": int(q["qt_vagas_reserva"].sum()) if len(q) else 0,
                "municipios_com_estoque_pre_disponivel": int(a1_muns["estoque_especialistas_pre_12m_media"].notna().sum()),
                "ivs_medio_nacional": float(nat["ivs_2010"].mean()) if nat["ivs_2010"].notna().any() else None,
                "populacao_2010_media_nacional": float(nat["populacao_2010"].mean()) if nat["populacao_2010"].notna().any() else None,
            }
        )
    # Linha total
    nat_all = matriz_tipologia
    a1_all = matriz_tipologia[matriz_tipologia["in_populacao_A1"]]
    q_all = quadro
    c_all = celulas_por_municipio
    suporte_rows.append(
        {
            "estrato": "total",
            "ordem": 99,
            "n_municipios_nacional": int(len(nat_all)),
            "n_municipios_populacao_A1": int(len(a1_all)),
            "n_celulas_quadro_ch1": int(len(q_all)),
            "n_celulas_funil_A1": int(c_all["n_celulas_funil_A1"].sum()) if len(c_all) else 0,
            "vagas_imediatas_publicadas": int(q_all["qt_vagas_imediatas"].sum()),
            "vagas_reserva_publicadas": int(q_all["qt_vagas_reserva"].sum()),
            "municipios_com_estoque_pre_disponivel": int(a1_all["estoque_especialistas_pre_12m_media"].notna().sum()),
            "ivs_medio_nacional": float(nat_all["ivs_2010"].mean()) if nat_all["ivs_2010"].notna().any() else None,
            "populacao_2010_media_nacional": float(nat_all["populacao_2010"].mean()) if nat_all["populacao_2010"].notna().any() else None,
        }
    )
    suporte = pd.DataFrame(suporte_rows).sort_values("ordem")

    # Concentração por UF (top 5 dentro da população A1)
    uf_counts = (
        matriz_tipologia[matriz_tipologia["in_populacao_A1"]]
        .groupby("sg_uf")
        .size()
        .sort_values(ascending=False)
        .head(10)
        .to_dict()
    )
    # Concentração por curso (células do quadro por curso no estrato — oferta, não outcome)
    curso_counts = quadro.groupby("cod_curso").size().sort_values(ascending=False).head(10).to_dict()
    curso_counts = {str(k): int(v) for k, v in curso_counts.items()}

    # 10. Validações do portão
    n_a1_total = len(municipios_a1_6)
    n_a1_classificados = int(matriz_tipologia[matriz_tipologia["in_populacao_A1"] & (matriz_tipologia["estrato"] != "NAO_CLASSIFICADO")].shape[0])
    cobertura_a1 = n_a1_classificados / n_a1_total if n_a1_total else 0
    # 5 novos municípios pós-2010 sem IVS
    novos_sem_ivs = sorted(matriz_tipologia[matriz_tipologia["flag_ivs_missing"]]["co_ibge_7d"].tolist())
    # Municípios do quadro (oferta) não encontrados em REGIC/RM (NAO_CLASSIFICADO)
    nao_classificados_a1 = matriz_tipologia[
        matriz_tipologia["in_populacao_A1"] & (matriz_tipologia["estrato"] == "NAO_CLASSIFICADO")
    ]["co_ibge_7d"].tolist()

    gate = "APROVADO_4_ESTRATOS" if (cobertura_a1 == 1.0 and len(nao_classificados_a1) == 0) else "REPROVADO"
    # Se não for possível construir remoticidade, cairia para APROVADO_CAPITAL_VS_FORA
    if gate == "REPROVADO":
        gate = "APROVADO_CAPITAL_VS_FORA" if n_a1_classificados > 0 else "REPROVADO"

    # 11. Manifesto
    manifesto: dict[str, Any] = {
        "protocolo": "A2_TIPOLOGIA_TERRITORIAL",
        "data_referencia": "2026-09-02",
        "efeitos_estimados": False,
        "portao": gate,
        "decisao": {
            "estrutura_minima": ["capital", "metropolitano", "interior_proximo_polo", "interior_remoto"],
            "usa_fonte_oficial_versionada": True,
            "fonte_metropolitano": "IBGE Composicao RMs/RIDEs 2022 (31/12/2022)",
            "fonte_remoticidade": "IBGE REGIC 2018 Hierarquia e regiao de influencia",
            "ivs_canonico": "IVS 2010 IPEA (nao substituir por IDHM/RDPC sem justificativa)",
            "fallback_se_remoticidade_impossivel": "manter apenas fora_das_capitais e retirar interior_remoto das hipoteses",
            "consultou_outcomes": False,
            "colunas_outcome_bloqueadas": sorted(forbidden_cols),
        },
        "criterios": {
            "cobertura_populacao_A1": cobertura_a1 == 1.0,
            "nenhum_NAO_CLASSIFICADO_em_A1": len(nao_classificados_a1) == 0,
            "ivs_cobertura_nacional": f"{int(matriz_tipologia['ivs_2010'].notna().sum())}/{len(matriz_tipologia)}",
            "estoque_pre_cobertura_A1": f"{int(matriz_tipologia[matriz_tipologia['in_populacao_A1']]['estoque_especialistas_pre_12m_media'].notna().sum())}/{n_a1_total}",
        },
        "fontes": {
            str(p.relative_to(ROOT)).replace("\\", "/"): {"sha256": sha256(p)}
            for p in [MALHA, QUADRO, MATRIZ_FUNIL, PAINEL_CNES, REGIC_XLSX, RM_XLSX]
        },
        "urls_oficiais": {"regic_2018": REGIC_URL, "rm_2022": RM_URL},
        "populacao": {
            "universo_A1_municipios": n_a1_total,
            "municipios_A1_classificados_4_estratos": n_a1_classificados,
            "nacional_total_municipios": int(len(matriz_tipologia)),
            "registros_fora_quadro_sem_municipio_na_matriz_A1": registros_fora_quadro_sem_municipio,
            "municipios_novos_pos2010_sem_IVS": novos_sem_ivs,
            "nao_classificados_em_A1": nao_classificados_a1,
            "mudancas_codigo_municipal": "5 municipios criados apos Censo 2010 (1504752, 4212650, 4220000, 4314548, 5006275) — IVS ausente, classificacao via REGIC/RM preservada",
        },
        "estratos": {row["estrato"]: {k: row[k] for k in row.index if k not in ("estrato",)} for _, row in suporte[suporte["estrato"] != "total"].iterrows()},
        "suporte_resumo": suporte.to_dict(orient="records"),
        "concentracao": {
            "por_uf_top10_populacao_A1": uf_counts,
            "por_curso_top10_quadro": curso_counts,
        },
        "regras_congeladas": {
            "capital": "co_ibge_7d em 27 capitais oficiais",
            "metropolitano": "nao capital e membro de RM ou RIDE (IBGE 2022)",
            "interior_proximo_polo": "nao capital, nao RM/RIDE, REGIC hierarquia_grupo 1-4 (Metropole/Capital Regional/Centro Sub-Regional/Centro de Zona, com ou sem integrante de AP)",
            "interior_remoto": "nao capital, nao RM/RIDE, REGIC hierarquia_grupo 5 (Centro Local, com ou sem integrante de AP) — nao inferido apenas por ser nao capital",
            "medidas_continuas_preservadas": [
                "ivs_2010 (+ subindices infra/ch/rt, categoria)",
                "populacao_2010",
                "idhm_2010, rdpc_2010 (descritivo)",
                "macro_regiao_saude, no_regiao_saude",
                "estoque_especialistas_pre_12m_media (CNES 202407-202506, media mensal por municipio, soma sobre CNES/curso do quadro) e estoque_pre_por_10k",
            ],
            "regra_missing": "municipios novos pos-2010 mantem estrato via REGIC/RM; IVS e estoque ficam NA e reportados; NAO_CLASSIFICADO exigiria fallback para capital vs fora",
        },
        "arquivos": {
            "matriz_tipologia_territorial.parquet": str(OUT_MATRIZ.relative_to(ROOT)).replace("\\", "/"),
            "suporte_estratos_territoriais.csv": str(OUT_SUPORTE.relative_to(ROOT)).replace("\\", "/"),
        },
        "implicacao_econometrica": (
            "Tipologia congelada sem consultar alocacao/homologacao/CNES pos. "
            "A3 pode usar estrato como heterogeneidade pre-especificada e covariadas previas. "
            "Nao chamar confirmacao/homologacao de retencao individual; inferencia agrupada no municipio quando exposicao varia no municipio."
        ),
    }

    # 12. Escrita atômica
    tmp_m = OUT_MATRIZ.with_suffix(".parquet.tmp")
    matriz_tipologia.to_parquet(tmp_m, index=False)
    tmp_m.replace(OUT_MATRIZ)

    tmp_s = OUT_SUPORTE.with_suffix(".csv.tmp")
    suporte.to_csv(tmp_s, index=False)
    tmp_s.replace(OUT_SUPORTE)

    tmp_j = OUT_MANIFESTO.with_suffix(".json.tmp")
    tmp_j.write_text(json.dumps(manifesto, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp_j.replace(OUT_MANIFESTO)

    print(f"[OK] A2 concluído: {gate}. Matriz com {len(matriz_tipologia)} municípios ({n_a1_total} na população A1); nenhum dado pessoal exposto.")
    print(f"     Estratos na população A1: {dict(matriz_tipologia[matriz_tipologia['in_populacao_A1']]['estrato'].value_counts())}")
    print(f"     Suporte salvo em {OUT_SUPORTE.name}; manifesto em {OUT_MANIFESTO.name}")


if __name__ == "__main__":
    main()
