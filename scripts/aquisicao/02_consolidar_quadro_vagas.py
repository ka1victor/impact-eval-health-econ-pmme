"""02_consolidar_quadro_vagas.py — Consolidação de Vagas e Definição do Tratamento PMM-E.

Este script lê, valida e unifica todas as 19 planilhas de oferta de vagas, alocação e
homologação do PMM-E (Ciclos 1, 2 e 3) presentes em `data/raw/aquisicao/vagas/` e `data/raw/pmm_e/`.

Ele produz:
1. `output/aquisicao/quadro_vagas_consolidado.parquet`: Registro completo de todas as vagas/células anunciadas.
2. `output/aquisicao/quadro_vagas_tratamento.parquet`: Painel de células canônicas de tratamento do Ciclo 1,
   Chamada 1 (24/07/2025), contendo as variáveis de tratamento (`immediate_is`), modalidade (IMEDIATA,
   RESERVA, DUPLA), faixas de atração e flags de sobreposição de CBO para a avaliação causal.
"""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
RAW_VAGAS_DIR = ROOT / "data" / "raw" / "aquisicao" / "vagas"
RAW_PMM_DIR = ROOT / "data" / "raw" / "pmm_e"
OUTPUT_DIR = ROOT / "output" / "aquisicao"
PONTE_FILE = OUTPUT_DIR / "ponte_curso_cbo_oficial.json"

OUT_CONSOLIDADO = OUTPUT_DIR / "quadro_vagas_consolidado.parquet"
OUT_TRATAMENTO = OUTPUT_DIR / "quadro_vagas_tratamento.parquet"


def normalize_str(val: object) -> str:
    if val is None:
        return ""
    text = unicodedata.normalize("NFKD", str(val))
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text).strip().upper()


def normalize_cnes(val: object) -> str:
    if val is None:
        return ""
    digits = re.sub(r"\D", "", str(val))
    if not digits or int(digits) == 0:
        return ""
    return digits.zfill(7)


def normalize_ibge(val: object) -> str:
    if val is None:
        return ""
    digits = re.sub(r"\D", "", str(val))
    return digits[:6] if len(digits) >= 6 else digits


def extract_course_id(val: object) -> Optional[int]:
    norm = normalize_str(val)
    if not norm:
        return None
    # Check leading number "01.", "1 -", etc.
    match = re.match(r"^(\d{1,2})", norm)
    if match:
        cid = int(match.group(1))
        if 1 <= cid <= 16:
            return cid

    # Match by key phrases in course name
    course_keywords = [
        (1, "ANESTESIOLOGIA"),
        (2, "CIRURGIA GERAL MINIMAMENTE"),
        (3, "CIRURGIA ONCOLOGICA"),
        (4, "COLOPROCTOLOGICA"),
        (5, "CIRURGIA DO APARELHO DIGESTIVO"),
        (6, "CIRURGIA GINECOLOGICA"),
        (7, "COLONOSCOPIA"),
        (8, "COLPOSCOPIA"),
        (9, "ECOCARDIOGRAFIA"),
        (10, "ENDOSCOPIA DIGESTIVA AVANCADA"),
        (11, "ENDOSCOPIA DIGESTIVA: ALTA"),
        (12, "ONCOLOGIA CLINICA"),
        (13, "RADIOTERAPIA"),
        (14, "ULTRASSONOGRAFIA MAMARIA"),
        (15, "VIDEOLARINGOSCOPIA"),
        (16, "ANATOMIA PATOLOGICA"),
    ]
    for cid, kw in course_keywords:
        if kw in norm:
            return cid
    return None


def parse_numeric(val: object) -> int:
    if val is None:
        return 0
    try:
        num = float(str(val).replace(",", ".").strip())
        return max(0, int(round(num)))
    except (ValueError, TypeError):
        return 0


# Mapeamento oficial dos arquivos de vagas e suas configurações
SPECS = [
    # CICLO 1
    {
        "file": RAW_VAGAS_DIR / "2025_ciclo1_chamada1_vagas.xlsx",
        "ciclo": 1,
        "chamada": 1,
        "tipo": "OFERTA_VAGAS",
        "versao_edital": "original",
        "dt_publicacao": "2025-07-24",
        "skiprows_min": 1,
        "col_imediata_idx": 10,
        "col_reserva_idx": 14,
    },
    {
        "file": RAW_VAGAS_DIR / "2025_ciclo1_chamada1_alocacao_retificada.xlsx",
        "ciclo": 1,
        "chamada": 1,
        "tipo": "ALOCACAO",
        "versao_edital": "retificada",
        "dt_publicacao": "2025-09-10",
        "skiprows_min": 0,
    },
    {
        "file": RAW_VAGAS_DIR / "2025_ciclo1_chamada1_alocacao_retificada_subjudice.xlsx",
        "ciclo": 1,
        "chamada": 1,
        "tipo": "ALOCACAO",
        "versao_edital": "retificada_subjudice",
        "dt_publicacao": "2025-09-19",
        "skiprows_min": 0,
    },
    {
        "file": RAW_VAGAS_DIR / "2025_ciclo1_chamada1_realocacao_retificado.xlsx",
        "ciclo": 1,
        "chamada": 1,
        "tipo": "REALOCACAO",
        "versao_edital": "retificada",
        "dt_publicacao": "2025-09-10",
        "skiprows_min": 0,
    },
    {
        "file": RAW_PMM_DIR / "2025_ciclo1_chamada1_homologados.xlsx",
        "ciclo": 1,
        "chamada": 1,
        "tipo": "HOMOLOGADOS",
        "versao_edital": "retificada",
        "dt_publicacao": "2025-09-29",
        "skiprows_min": 0,
    },
    {
        "file": RAW_PMM_DIR / "2025_ciclo1_chamada2_vagas_e_alocados.xlsx",
        "ciclo": 1,
        "chamada": 2,
        "tipo": "OFERTA_E_ALOCADOS",
        "versao_edital": "oficial",
        "dt_publicacao": "2025-09-29",
        "skiprows_min": 0,
    },
    {
        "file": RAW_PMM_DIR / "2025_ciclo1_chamada2_classificacao_final.xlsx",
        "ciclo": 1,
        "chamada": 2,
        "tipo": "CLASSIFICACAO",
        "versao_edital": "final",
        "dt_publicacao": "2025-11-14",
        "skiprows_min": 1,
    },
    {
        "file": RAW_PMM_DIR / "2025_ciclo1_chamada2_homologados.xlsx",
        "ciclo": 1,
        "chamada": 2,
        "tipo": "HOMOLOGADOS",
        "versao_edital": "final",
        "dt_publicacao": "2025-11-24",
        "skiprows_min": 0,
    },
    # CICLO 2
    {
        "file": RAW_VAGAS_DIR / "2026_ciclo2_chamada1_vagas_e_servicos_original.xlsx",
        "ciclo": 2,
        "chamada": 1,
        "tipo": "OFERTA_VAGAS",
        "versao_edital": "original",
        "dt_publicacao": "2026-02-03",
        "skiprows_min": 10,
        "col_imediata_idx": 10,
        "col_reserva_idx": 14,
    },
    {
        "file": RAW_VAGAS_DIR / "2026_ciclo2_chamada1_vagas_e_servicos_retificado.xlsx",
        "ciclo": 2,
        "chamada": 1,
        "tipo": "OFERTA_VAGAS",
        "versao_edital": "retificada_servicos",
        "dt_publicacao": "2026-02-13",
        "skiprows_min": 1,
        "col_imediata_idx": 10,
        "col_reserva_idx": 14,
    },
    {
        "file": RAW_PMM_DIR / "2026_ciclo2_chamada1_vagas_retificadas.xlsx",
        "ciclo": 2,
        "chamada": 1,
        "tipo": "OFERTA_VAGAS",
        "versao_edital": "retificada_final",
        "dt_publicacao": "2026-03-19",
        "skiprows_min": 1,
        "col_imediata_idx": 10,
        "col_reserva_idx": 14,
    },
    {
        "file": RAW_PMM_DIR / "2026_ciclo2_chamada1_resultado_final_remanescentes.xlsx",
        "ciclo": 2,
        "chamada": 1,
        "tipo": "RESULTADO_FINAL",
        "versao_edital": "final",
        "dt_publicacao": "2026-05-05",
        "skiprows_min": 0,
    },
    {
        "file": RAW_PMM_DIR / "2026_ciclo2_chamada2_vagas.xlsx",
        "ciclo": 2,
        "chamada": 2,
        "tipo": "OFERTA_VAGAS",
        "versao_edital": "oficial",
        "dt_publicacao": "2026-04-16",
        "skiprows_min": 1,
        "col_imediata_idx": 10,
        "col_reserva_idx": 14,
    },
    {
        "file": RAW_PMM_DIR / "2026_ciclo2_chamada2_resultado_final.xlsx",
        "ciclo": 2,
        "chamada": 2,
        "tipo": "RESULTADO_FINAL",
        "versao_edital": "final",
        "dt_publicacao": "2026-05-28",
        "skiprows_min": 0,
    },
    # CICLO 3
    {
        "file": RAW_VAGAS_DIR / "2026_ciclo3_gestores_quadro_vagas_original.xlsx",
        "ciclo": 3,
        "chamada": 0,
        "tipo": "ADESAO_GESTORES",
        "versao_edital": "original",
        "dt_publicacao": "2026-05-15",
        "skiprows_min": 0,
    },
    {
        "file": RAW_PMM_DIR / "2026_ciclo3_adesao_gestores_resultado_final.xlsx",
        "ciclo": 3,
        "chamada": 0,
        "tipo": "ADESAO_GESTORES",
        "versao_edital": "final",
        "dt_publicacao": "2026-07-15",
        "skiprows_min": 0,
    },
    {
        "file": RAW_VAGAS_DIR / "2026_ciclo3_chamada1_vagas_original.xlsx",
        "ciclo": 3,
        "chamada": 1,
        "tipo": "OFERTA_VAGAS",
        "versao_edital": "original",
        "dt_publicacao": "2026-07-16",
        "skiprows_min": 2,
        "col_imediata_idx": 11,
        "col_reserva_idx": 15,
    },
    {
        "file": RAW_PMM_DIR / "2026_ciclo3_chamada1_vagas_retificadas.xlsx",
        "ciclo": 3,
        "chamada": 1,
        "tipo": "OFERTA_VAGAS",
        "versao_edital": "retificada",
        "dt_publicacao": "2026-07-24",
        "skiprows_min": 2,
        "col_imediata_idx": 11,
        "col_reserva_idx": 15,
    },
    {
        "file": RAW_PMM_DIR / "2026_ciclo3_chamada1_resultado_final_sub_judice.xlsx",
        "ciclo": 3,
        "chamada": 1,
        "tipo": "RESULTADO_FINAL",
        "versao_edital": "final_sub_judice",
        "dt_publicacao": "2026-08-25",
        "skiprows_min": 0,
    },
]


def processar_planilhas() -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Processa todas as 19 planilhas e extrai o painel consolidado e o painel de tratamento."""
    ponte_data = {}
    if PONTE_FILE.exists():
        with PONTE_FILE.open("r", encoding="utf-8") as f:
            ponte_json = json.load(f)
            ponte_data = {c["cod_curso"]: c for c in ponte_json.get("catalogo_cursos", [])}

    consolidated_records: List[Dict[str, Any]] = []

    for spec in SPECS:
        path = spec["file"]
        if not path.exists():
            print(f"[AVISO] Arquivo não encontrado: {path.name}")
            continue

        print(f"[Vagas] Lendo {path.name}...")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Localizar linha de cabeçalho
            header_idx = -1
            for idx, r in enumerate(rows[:30]):
                norm_r = [normalize_str(c) for c in r if c is not None]
                if any("CNES" in c for c in norm_r) and any(
                    "CURSO" in c or "APRIMORAMENTO" in c or "TIPO DO ESTABELECIMENTO" in c
                    for c in norm_r
                ):
                    header_idx = idx
                    break

            if header_idx == -1:
                continue

            header_raw = [normalize_str(c) for c in rows[header_idx]]
            data_rows = rows[header_idx + 1 :]

            # Localizar índices de colunas
            col_cnes = next((i for i, c in enumerate(header_raw) if "CNES" in c), None)
            col_curso = next(
                (i for i, c in enumerate(header_raw) if "CURSO" in c or "APRIMORAMENTO" in c),
                None,
            )
            col_ibge = next((i for i, c in enumerate(header_raw) if "IBGE" in c), None)
            col_uf = next((i for i, c in enumerate(header_raw) if c == "UF"), None)
            col_mun = next((i for i, c in enumerate(header_raw) if "MUNIC" in c), None)
            col_estab = next(
                (
                    i
                    for i, c in enumerate(header_raw)
                    if "ESTABELECIMENTO" in c or "NOME FANTASIA" in c
                ),
                None,
            )
            col_faixa = next((i for i, c in enumerate(header_raw) if "FAIXA" in c), None)
            col_gestao = next((i for i, c in enumerate(header_raw) if "GESTAO" in c), None)

            # Índices de vagas imediatas e reserva se disponíveis
            col_im = spec.get("col_imediata_idx")
            col_res = spec.get("col_reserva_idx")

            # Se for Ciclo 1 Chamada 2 Reserva
            is_c1_ch2_reserva = (
                spec["ciclo"] == 1
                and spec["chamada"] == 2
                and "RESERVA" in normalize_str(sheet_name)
            )

            for r in data_rows:
                if not r or not any(r):
                    continue
                cnes = normalize_cnes(r[col_cnes]) if col_cnes is not None and col_cnes < len(r) else ""
                if not cnes:
                    continue

                cid = (
                    extract_course_id(r[col_curso])
                    if col_curso is not None and col_curso < len(r)
                    else None
                )
                raw_curso_name = str(r[col_curso]) if col_curso is not None and col_curso < len(r) else ""

                ibge = normalize_ibge(r[col_ibge]) if col_ibge is not None and col_ibge < len(r) else ""
                uf = normalize_str(r[col_uf]) if col_uf is not None and col_uf < len(r) else ""
                mun = normalize_str(r[col_mun]) if col_mun is not None and col_mun < len(r) else ""
                estab = normalize_str(r[col_estab]) if col_estab is not None and col_estab < len(r) else ""
                faixa = normalize_str(r[col_faixa]) if col_faixa is not None and col_faixa < len(r) else ""
                gestao = normalize_str(r[col_gestao]) if col_gestao is not None and col_gestao < len(r) else ""

                v_im = 0
                v_res = 0

                if col_im is not None and col_im < len(r):
                    v_im = parse_numeric(r[col_im])
                if col_res is not None and col_res < len(r):
                    v_res = parse_numeric(r[col_res])

                if is_c1_ch2_reserva:
                    v_res = parse_numeric(r[9]) if len(r) > 9 else 0
                    v_im = 0

                # Determinar modalidade
                if v_im > 0 and v_res > 0:
                    modalidade = "DUPLA"
                elif v_im > 0:
                    modalidade = "IMEDIATA"
                elif v_res > 0:
                    modalidade = "RESERVA"
                else:
                    modalidade = "NAO_ESPECIFICADA"

                no_curso_padrao = ponte_data.get(cid, {}).get("no_curso_padronizado", raw_curso_name) if cid else raw_curso_name

                consolidated_records.append(
                    {
                        "co_cnes_7d": cnes,
                        "cod_curso": cid,
                        "no_curso": no_curso_padrao,
                        "raw_curso": raw_curso_name,
                        "ciclo": spec["ciclo"],
                        "chamada": spec["chamada"],
                        "tipo_documento": spec["tipo"],
                        "versao_edital": spec["versao_edital"],
                        "dt_publicacao": spec["dt_publicacao"],
                        "sheet_name": sheet_name,
                        "modalidade_vaga": modalidade,
                        "qt_vagas_imediatas": v_im,
                        "qt_vagas_reserva": v_res,
                        "qt_vagas_total": v_im + v_res,
                        "faixa_atracao_anunciada": faixa,
                        "co_ibge_6d": ibge,
                        "sg_uf": uf,
                        "no_municipio": mun,
                        "no_estabelecimento": estab,
                        "tipo_gestao": gestao,
                        "arquivo_origem": path.name,
                    }
                )

    df_consolidado = pd.DataFrame(consolidated_records)

    # ----------------------------------------------------
    # Construção do Quadro Canônico de Tratamento (Ciclo 1 Chamada 1)
    # ----------------------------------------------------
    # Filtro exato na publicação inicial de 24/07/2025
    mask_c1_ch1 = (
        (df_consolidado["ciclo"] == 1)
        & (df_consolidado["chamada"] == 1)
        & (df_consolidado["tipo_documento"] == "OFERTA_VAGAS")
        & (df_consolidado["versao_edital"] == "original")
        & (df_consolidado["cod_curso"].notna())
    )
    df_c1 = df_consolidado[mask_c1_ch1].copy()
    df_c1["cod_curso"] = df_c1["cod_curso"].astype(int)

    # Agregação por Célula CNES-Curso no Ciclo 1 Chamada 1
    cell_grouped = (
        df_c1.groupby(["co_cnes_7d", "cod_curso"])
        .agg(
            {
                "no_curso": "first",
                "qt_vagas_imediatas": "sum",
                "qt_vagas_reserva": "sum",
                "qt_vagas_total": "sum",
                "faixa_atracao_anunciada": "first",
                "co_ibge_6d": "first",
                "sg_uf": "first",
                "no_municipio": "first",
                "no_estabelecimento": "first",
                "tipo_gestao": "first",
                "dt_publicacao": "first",
            }
        )
        .reset_index()
    )

    def classify_treatment(row: pd.Series) -> str:
        im = row["qt_vagas_imediatas"]
        res = row["qt_vagas_reserva"]
        if im > 0 and res > 0:
            return "DUPLA"
        elif im > 0:
            return "IMEDIATA"
        elif res > 0:
            return "RESERVA"
        return "SEM_VAGAS"

    cell_grouped["modalidade_original"] = cell_grouped.apply(classify_treatment, axis=1)
    cell_grouped["immediate_is"] = (cell_grouped["modalidade_original"] == "IMEDIATA").astype(int)

    # Identificar sobreposição de CBO
    overlap_courses = [
        c["cod_curso"]
        for c in ponte_data.values()
        if c.get("sobreposicao", False)
    ]
    cell_grouped["flag_overlap_cbo"] = cell_grouped["cod_curso"].isin(overlap_courses)

    # Identificar amostra principal sem dupla modalidade e com contraste limpo
    cell_grouped["amostra_principal"] = cell_grouped["modalidade_original"].isin(["IMEDIATA", "RESERVA"])

    # Contagem de estabelecimentos com múltiplas modalidades
    cnes_modalidades = cell_grouped.groupby("co_cnes_7d")["modalidade_original"].nunique()
    multi_modal_cnes = set(cnes_modalidades[cnes_modalidades > 1].index)
    cell_grouped["flag_cnes_multiplas_modalidades"] = cell_grouped["co_cnes_7d"].isin(multi_modal_cnes)

    return df_consolidado, cell_grouped


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("=== [Subagente 1] Consolidando Quadros de Vagas e Definindo Matriz de Tratamento ===")

    df_consolidado, df_tratamento = processar_planilhas()

    # Salvar Parquets
    df_consolidado.to_parquet(OUT_CONSOLIDADO, index=False)
    df_tratamento.to_parquet(OUT_TRATAMENTO, index=False)

    print(f"\n[OK] Quadro consolidado salvo em: {OUT_CONSOLIDADO}")
    print(f"     Total de linhas processadas: {len(df_consolidado):,}")
    print(f"     Planilhas de origem: {df_consolidado['arquivo_origem'].nunique()} arquivos")

    print(f"\n[OK] Quadro canônico de tratamento salvo em: {OUT_TRATAMENTO}")
    print(f"     Total de células (CNES x Curso): {len(df_tratamento):,}")
    print(f"     Distribuição de modalidade:")
    for mod, count in df_tratamento["modalidade_original"].value_counts().items():
        print(f"       - {mod}: {count} células")
    print(f"     Células na amostra principal (Imediata vs Reserva): {df_tratamento['amostra_principal'].sum()}")
    print(f"     CNES com ambas as modalidades (identificação within-CNES): {df_tratamento['flag_cnes_multiplas_modalidades'].sum()}")


if __name__ == "__main__":
    main()
